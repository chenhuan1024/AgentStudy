import csv
import asyncio
import math
import os
import re
import sys
import threading
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple

import duckdb
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook


if getattr(sys, "frozen", False):
    BUNDLE_DIR = Path(getattr(sys, "_MEIPASS"))
    RUNTIME_DIR = Path(sys.executable).resolve().parent
    FRONTEND_DIR = BUNDLE_DIR / "frontend_build"
else:
    PROJECT_DIR = Path(__file__).resolve().parents[2]
    RUNTIME_DIR = PROJECT_DIR / "backend"
    FRONTEND_DIR = PROJECT_DIR / "frontend" / "build"

DATA_DIR = RUNTIME_DIR / "data"
EXPORT_DIR = RUNTIME_DIR / "exports"
DB_FILE = RUNTIME_DIR / "site_selection.duckdb"

DATA_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

MAX_DEFAULT_ROWS = 1000
ALLOWED_OPERATORS = {">", "<", ">=", "<=", "="}
ALLOWED_EXPORT_FORMATS = {"csv", "excel"}
ROLE_TABLE_MAP = {
    "engineering": "engineering_view",
    "traffic": "traffic_view",
    "station": "station_view",
}


def safe_identifier(name: str) -> str:
    if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
        raise HTTPException(status_code=400, detail=f"非法标识符: {name}")
    return name


def sql_quote_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_quote_ident(name: str) -> str:
    # DuckDB 标识符用双引号包裹；内部双引号需要转义为两个双引号
    return '"' + name.replace('"', '""') + '"'


def safe_filename(name: str) -> str:
    # Windows 文件名非法字符: \ / : * ? " < > |
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned[:120] if cleaned else "result"


def safe_excel_sheet_title(name: str) -> str:
    # Excel 工作表名最多 31 字符，且不能包含 : \ / ? * [ ]
    raw = (name or "").strip() or "Sheet"
    cleaned = re.sub(r'[:\\/?*\[\]]+', "_", raw)
    return cleaned[:31] if cleaned else "Sheet"


def sql_literal(value: Any) -> str:
    # 将 Python 值安全转成 SQL 字面量（用于不支持 prepared params 的 DDL 语句）
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        # DuckDB 接受标准数字文本
        return str(value)
    # 其他一律按字符串处理（包含中文、特殊字符）
    return "'" + str(value).replace("'", "''") + "'"


def pretty_num(value: float) -> str:
    if value is None:
        return ""
    if abs(value) >= 1000:
        return f"{value:.2f}".rstrip("0").rstrip(".")
    if abs(value) >= 1:
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return f"{value:.8f}".rstrip("0").rstrip(".")


class FilterCondition(BaseModel):
    field: str
    operator: Literal[">", "<", ">=", "<=", "="]
    value: Any


class PreviewColumnFilter(BaseModel):
    """工参/话务预览：列头筛选（列内 OR，列间 AND）"""

    field: str
    values: List[Any] = Field(default_factory=list)


class PreviewRequest(BaseModel):
    role: Literal["engineering", "traffic", "station"]
    limit: int = Field(default=100, ge=1, le=MAX_DEFAULT_ROWS)
    column_filters: List[PreviewColumnFilter] = Field(default_factory=list)


class ColumnDistinctRequest(BaseModel):
    role: Literal["engineering", "traffic", "station"]
    field: str
    max_values: int = Field(default=50000, ge=1, le=200000)


class ExportFilteredRequest(BaseModel):
    """按预览列筛选条件导出整张表（不限 100 行）"""

    role: Literal["engineering", "traffic", "station"]
    column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    file_format: Literal["csv", "excel"]


class FilterRequest(BaseModel):
    role: Literal["engineering", "traffic", "station"] = "engineering"
    conditions: List[FilterCondition] = []
    sort_field: Optional[str] = None
    sort_order: Optional[Literal["asc", "desc"]] = None
    limit: int = Field(default=MAX_DEFAULT_ROWS, ge=1, le=MAX_DEFAULT_ROWS)


class NlFilterRequest(BaseModel):
    text: Optional[str] = None
    condition_groups: Optional[List[List[FilterCondition]]] = None
    group_names: Optional[List[str]] = None
    traffic_role: Literal["traffic"] = "traffic"
    engineering_role: Literal["engineering"] = "engineering"
    station_role: Literal["station"] = "station"
    condition_role: Literal["traffic", "engineering", "station"] = "traffic"
    traffic_id_field: str = "cell_id"
    engineering_id_field: str = "cell_id"
    station_id_field: str = "cell_id"
    limit: int = Field(default=MAX_DEFAULT_ROWS, ge=1, le=MAX_DEFAULT_ROWS)
    # 与右侧预览列筛选一致：先缩小话务/工参范围，再套条件组
    traffic_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    engineering_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    station_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    # 当 condition_role=station 且提供此 key 时，条件作用于“选站关联话务预览”结果
    station_traffic_view_key: Optional[str] = None


class NlParseRequest(BaseModel):
    text: str
    traffic_role: Literal["traffic"] = "traffic"


class FieldMaxRequest(BaseModel):
    role: Literal["engineering", "traffic", "station"]
    field: str


class PasteUploadRequest(BaseModel):
    role: Literal["engineering", "traffic", "station"]
    content: str


class ExportRequest(BaseModel):
    result_key: str
    table_type: Literal["traffic", "engineering"]
    file_format: Literal["csv", "excel"]


class ResultPreviewRequest(BaseModel):
    result_key: str
    table_type: Literal["traffic", "engineering", "station"]
    limit: int = Field(default=100, ge=1, le=MAX_DEFAULT_ROWS)


class StationTrafficPreviewRequest(BaseModel):
    """选站关联话务预览（按 ID 关联，返回话务表）"""

    limit: int = Field(default=100, ge=1, le=MAX_DEFAULT_ROWS)
    traffic_id_field: str = "cell_id"
    station_id_field: str = "cell_id"
    traffic_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    station_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    view_key: Optional[str] = None


class StationTrafficDistinctRequest(BaseModel):
    view_key: str
    field: str
    max_values: int = Field(default=50000, ge=1, le=200000)
    column_filters: List[PreviewColumnFilter] = Field(default_factory=list)


class StationDistinctCountRequest(BaseModel):
    field: str


class ConditionChartCountRequest(BaseModel):
    """条件组图表统计：返回筛选前后数量（总组 + 每个单条件）"""

    condition_role: Literal["traffic", "engineering", "station"] = "traffic"
    conditions: List[FilterCondition] = Field(default_factory=list)
    traffic_id_field: str = "cell_id"
    engineering_id_field: str = "cell_id"
    station_id_field: str = "cell_id"
    traffic_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    engineering_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    station_column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    station_traffic_view_key: Optional[str] = None


class ExportNlBatchItem(BaseModel):
    result_key: str
    group_name: str = ""


class ExportNlBatchRequest(BaseModel):
    """一键导出所有条件组（同一文件：Excel 多工作表 或 CSV 合并并带条件组列）"""

    groups: List[ExportNlBatchItem]
    table_type: Literal["traffic", "engineering"]
    file_format: Literal["csv", "excel"]


class EngineeringChartRequest(BaseModel):
    """工参表列分布 / 阈值占比（可叠加预览列筛选）"""

    field: str
    column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    mode: Literal["top_values", "threshold"] = "top_values"
    top_n: int = Field(default=20, ge=1, le=200)
    operator: Optional[Literal[">", "<", ">=", "<=", "="]] = None
    threshold_value: Optional[float] = None


class ColumnDistributionRequest(BaseModel):
    """通用列分布：自动识别类型；数值支持直方图/阈值三段"""

    role: Literal["engineering", "traffic", "station"]
    field: str
    column_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    mode: Literal["auto", "histogram", "threshold_3bins"] = "auto"
    top_n: int = Field(default=20, ge=1, le=200)
    bins: int = Field(default=12, ge=3, le=60)
    x1: Optional[float] = None
    x2: Optional[float] = None


class CompareDistributionRequest(BaseModel):
    """全量 vs 条件过滤后的双分布对比"""

    table_name: str
    field: str
    condition_list: List[FilterCondition] = Field(default_factory=list)
    mode: Literal["auto", "top_values", "histogram"] = "auto"
    top_n: int = Field(default=20, ge=1, le=200)
    bins: int = Field(default=12, ge=3, le=60)


class UnifiedDistributionRequest(BaseModel):
    """统一分布接口：支持全量/条件过滤双层对比"""

    table_name: str
    column: str
    conditions: List[FilterCondition] = Field(default_factory=list)
    preview_filters: List[PreviewColumnFilter] = Field(default_factory=list)
    mode: Literal["auto", "top_values", "histogram", "threshold_3bins"] = "auto"
    bins: int = Field(default=12, ge=3, le=200)
    top_n: int = Field(default=20, ge=1, le=200)
    x1: Optional[float] = None
    x2: Optional[float] = None
    compare_with_base: bool = True
    # 直方图区间宽度（可选）；若设置则按步长划分桶数，上限 200
    bin_width: Optional[float] = Field(default=None, gt=0)


app = FastAPI(title="Site Selection Tool API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

con = duckdb.connect(database=str(DB_FILE), read_only=False)
CON_LOCK = threading.RLock()
RUNTIME_STARTED_AT = time.time()
REQUEST_SERIALIZER = asyncio.Lock()


@app.middleware("http")
async def _serialize_db_requests(request, call_next):
    # 单连接 DuckDB 在并发请求下容易触发结果集异常，这里串行化请求以提升稳定性
    async with REQUEST_SERIALIZER:
        return await call_next(request)

# 用于保存上传文件和结果查询定义，避免前端重复传递复杂 SQL
uploaded_meta: Dict[str, Dict[str, str]] = {}
result_queries: Dict[str, Dict[str, str]] = {}
# 选站关联话务预览临时视图 key -> view 名
station_traffic_views: Dict[str, str] = {}


def get_table_name_by_role(role: str) -> str:
    if role not in ROLE_TABLE_MAP:
        raise HTTPException(status_code=400, detail=f"未知角色: {role}")
    return ROLE_TABLE_MAP[role]


def ensure_table_exists(role: str) -> str:
    table_name = get_table_name_by_role(role)
    if role not in uploaded_meta:
        raise HTTPException(status_code=400, detail=f"{role} 文件尚未上传")
    return table_name


def resolve_table_name_input(table_name: str) -> str:
    """兼容 role 名或真实 DuckDB 视图名输入"""
    if table_name in ROLE_TABLE_MAP:
        return ensure_table_exists(table_name)
    safe = safe_identifier(table_name)
    try:
        con.execute(f"PRAGMA table_info({safe})").fetchall()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"表不存在: {table_name}") from exc
    return safe


def get_columns(table_name: str) -> List[str]:
    table_name = safe_identifier(table_name)
    for _ in range(2):
        try:
            with CON_LOCK:
                rows = con.execute(f"PRAGMA table_info({table_name})").fetchall()
            return [r[1] for r in rows]
        except duckdb.InvalidInputException as exc:
            if "No open result set" not in str(exc):
                raise
            time.sleep(0.03)
    raise HTTPException(status_code=500, detail="读取表结构失败，请重试")


def get_column_types(table_name: str) -> Dict[str, str]:
    table_name = safe_identifier(table_name)
    for _ in range(2):
        try:
            with CON_LOCK:
                rows = con.execute(f"PRAGMA table_info({table_name})").fetchall()
            # PRAGMA table_info: (cid, name, type, notnull, dflt_value, pk)
            return {r[1]: (r[2] or "") for r in rows}
        except duckdb.InvalidInputException as exc:
            if "No open result set" not in str(exc):
                raise
            time.sleep(0.03)
    raise HTTPException(status_code=500, detail="读取字段类型失败，请重试")


def build_where_clause(
    table_name: str,
    conditions: List[FilterCondition],
) -> Tuple[str, List[Any]]:
    columns = get_columns(table_name)
    column_map = {c.lower(): c for c in columns}
    col_types = get_column_types(table_name)
    where_parts: List[str] = []
    params: List[Any] = []
    for cond in conditions:
        if cond.operator not in ALLOWED_OPERATORS:
            raise HTTPException(status_code=400, detail=f"不支持操作符: {cond.operator}")
        real_field = column_map.get(cond.field.lower())
        if not real_field:
            raise HTTPException(status_code=400, detail=f"字段不存在: {cond.field}")

        col_type = (col_types.get(real_field) or "").upper()
        op = cond.operator
        val = cond.value

        # 解决 DuckDB Binder Error: VARCHAR 与数值比较需要显式 CAST
        # 若列是字符串但用户给了数值，使用 TRY_CAST(列 AS DOUBLE) 做数值比较（失败则为 NULL，不满足条件）
        is_numeric_val = isinstance(val, (int, float)) and op in {">", "<", ">=", "<="}
        is_string_col = any(t in col_type for t in ["VARCHAR", "TEXT", "STRING"])

        if is_string_col and is_numeric_val:
            where_parts.append(f"try_cast({sql_quote_ident(real_field)} AS DOUBLE) {op} ?")
            params.append(val)
        else:
            where_parts.append(f"{sql_quote_ident(real_field)} {op} ?")
            params.append(val)
    if not where_parts:
        return "", []
    return " WHERE " + " AND ".join(where_parts), params


def build_preview_column_filters_literal(
    table_name: str,
    filters: List[PreviewColumnFilter],
) -> str:
    """预览列筛选：生成无占位符的 WHERE，用于需要内联 SQL 的场景"""
    if not filters:
        return ""
    columns = get_columns(table_name)
    column_map = {c.lower(): c for c in columns}
    where_parts: List[str] = []
    for filt in filters:
        if not filt.values:
            continue
        real_field = column_map.get(filt.field.lower())
        if not real_field:
            raise HTTPException(status_code=400, detail=f"字段不存在: {filt.field}")
        col_sql = sql_quote_ident(real_field)
        vals = filt.values
        if len(vals) == 1:
            where_parts.append(f"{col_sql} = {sql_literal(vals[0])}")
        else:
            inner = ",".join(sql_literal(v) for v in vals)
            where_parts.append(f"{col_sql} IN ({inner})")
    if not where_parts:
        return ""
    return " WHERE " + " AND ".join(where_parts)


def build_where_clause_literal(table_name: str, conditions: List[FilterCondition]) -> str:
    # 生成不含 prepared 参数的 WHERE，用于 CREATE VIEW/TABLE AS 这类语句
    columns = get_columns(table_name)
    column_map = {c.lower(): c for c in columns}
    col_types = get_column_types(table_name)
    where_parts: List[str] = []

    for cond in conditions:
        if cond.operator not in ALLOWED_OPERATORS:
            raise HTTPException(status_code=400, detail=f"不支持操作符: {cond.operator}")
        real_field = column_map.get(cond.field.lower())
        if not real_field:
            raise HTTPException(status_code=400, detail=f"字段不存在: {cond.field}")

        col_type = (col_types.get(real_field) or "").upper()
        op = cond.operator
        val = cond.value

        is_numeric_val = isinstance(val, (int, float)) and op in {">", "<", ">=", "<="}
        is_string_col = any(t in col_type for t in ["VARCHAR", "TEXT", "STRING"])

        if is_string_col and is_numeric_val:
            where_parts.append(f"try_cast({sql_quote_ident(real_field)} AS DOUBLE) {op} {sql_literal(val)}")
        else:
            where_parts.append(f"{sql_quote_ident(real_field)} {op} {sql_literal(val)}")

    if not where_parts:
        return ""
    return " WHERE " + " AND ".join(where_parts)


def resolve_real_column(table_name: str, field_name: str) -> str:
    column_map = {c.lower(): c for c in get_columns(table_name)}
    real_name = column_map.get(field_name.lower())
    if not real_name:
        raise HTTPException(status_code=400, detail=f"字段不存在: {field_name}")
    return real_name


# Excel/导入常见无效占位（统一用小写与 trim 后比较）
_INVALID_STR_LABELS_LOWER: Tuple[str, ...] = (
    "!value",
    "#value!",
    "#n/a",
    "n/a",
    "na",
    "#ref!",
    "-",
    "null",
)


def _junk_where_varchar(col_sql: str) -> str:
    literals = ",".join(sql_literal(s) for s in _INVALID_STR_LABELS_LOWER)
    expr = f"lower(trim(CAST(t.{col_sql} AS VARCHAR)))"
    return f"({expr} NOT IN ({literals}))"


def _finite_double_expr(col_sql: str) -> str:
    return f"try_cast(t.{col_sql} AS DOUBLE)"


def _finite_double_where(col_sql: str) -> str:
    v = _finite_double_expr(col_sql)
    return f"({v} IS NOT NULL AND isfinite({v}))"


def _numeric_parse_ratio(base_sub: str, col_sql: str) -> float:
    v = _finite_double_expr(col_sql)
    row = con.execute(
        f"""
        SELECT coalesce(avg(CASE WHEN {v} IS NOT NULL AND isfinite({v}) THEN 1.0 ELSE 0.0 END), 0.0)
        FROM ({base_sub}) t
        """
    ).fetchone()
    return float(row[0]) if row and row[0] is not None else 0.0


def build_preview_column_filters_where(
    table_name: str,
    filters: List[PreviewColumnFilter],
) -> Tuple[str, List[Any]]:
    if not filters:
        return "", []
    columns = get_columns(table_name)
    column_map = {c.lower(): c for c in columns}
    where_parts: List[str] = []
    params: List[Any] = []
    for filt in filters:
        if not filt.values:
            continue
        real_field = column_map.get(filt.field.lower())
        if not real_field:
            raise HTTPException(status_code=400, detail=f"字段不存在: {filt.field}")
        col_sql = sql_quote_ident(real_field)
        vals = filt.values
        if len(vals) == 1:
            where_parts.append(f"{col_sql} = ?")
            params.append(vals[0])
        else:
            ph = ",".join(["?"] * len(vals))
            where_parts.append(f"{col_sql} IN ({ph})")
            params.extend(vals)
    if not where_parts:
        return "", []
    return " WHERE " + " AND ".join(where_parts), params


def execute_preview_query(
    table_name: str,
    limit: int,
    column_filters: Optional[List[PreviewColumnFilter]] = None,
) -> Dict[str, Any]:
    table_name = safe_identifier(table_name)
    limit = min(limit, MAX_DEFAULT_ROWS)
    where_clause, params = build_preview_column_filters_where(
        table_name, column_filters or []
    )
    query = f"SELECT * FROM {table_name}{where_clause} LIMIT {limit}"
    cursor = con.execute(query, params)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]
    # 原始表总行数（不随列筛选变化，供前端「Total」锁定展示）
    try:
        total_row_count = int(con.execute(f"SELECT count(*) FROM {table_name}").fetchone()[0])
    except Exception:
        total_row_count = 0
    matching_count: Optional[int] = None
    # 无列筛选时，筛选后行数必然等于总行数，避免重复全表 count。
    if not where_clause:
        matching_count = total_row_count
    else:
        try:
            cnt_row = con.execute(f"SELECT count(*) FROM {table_name}{where_clause}", params).fetchone()
            matching_count = int(cnt_row[0]) if cnt_row else 0
        except Exception:
            matching_count = None
    return {
        "columns": columns,
        "rows": data,
        "total_row_count": total_row_count,
        "matching_row_count": matching_count,
    }


def parse_nl_conditions(text: str) -> List[List[FilterCondition]]:
    groups: List[List[FilterCondition]] = []
    raw_groups = [g.strip() for g in re.split(r"[;；\n]+", text) if g.strip()]
    if not raw_groups:
        raise HTTPException(status_code=400, detail="未解析到任何条件组")

    for raw_group in raw_groups:
        # 支持 "且 / and / AND / && / 逗号" 作为条件连接
        parts = [
            p.strip()
            for p in re.split(r"\s*(?:且|and|AND|&&|,|，)\s*", raw_group)
            if p.strip()
        ]
        conds: List[FilterCondition] = []
        for part in parts:
            match = re.match(
                r"^([A-Za-z_][A-Za-z0-9_]*)\s*(>=|<=|=|>|<)\s*(.+)$",
                part,
            )
            if not match:
                raise HTTPException(status_code=400, detail=f"条件无法解析: {part}")
            field, operator, raw_value = match.groups()
            raw_value = raw_value.strip()
            if (raw_value.startswith("'") and raw_value.endswith("'")) or (
                raw_value.startswith('"') and raw_value.endswith('"')
            ):
                value: Any = raw_value[1:-1]
            else:
                try:
                    value = int(raw_value)
                except ValueError:
                    try:
                        value = float(raw_value)
                    except ValueError:
                        value = raw_value
            conds.append(FilterCondition(field=field, operator=operator, value=value))
        groups.append(conds)
    return groups


def _cell_text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def detect_header_skip_rows(rows: List[List[Any]], max_skip: int = 6) -> int:
    """
    在前 max_skip+1 行中估计表头所在行，返回需跳过的行数。
    典型场景：首行即表头（skip=0）或前几行说明文字后第 7 行为表头（skip=6）。
    """
    if not rows:
        return 0
    upper = min(len(rows), max_skip + 1)
    for i in range(upper):
        current = rows[i]
        non_empty = [_cell_text(x) for x in current if _cell_text(x)]
        if len(non_empty) < 2:
            continue
        uniq_ratio = len(set(non_empty)) / max(1, len(non_empty))
        if uniq_ratio < 0.6:
            continue
        # 表头之后至少要有一行像数据
        if i + 1 < len(rows):
            nxt_non_empty = [_cell_text(x) for x in rows[i + 1] if _cell_text(x)]
            if len(nxt_non_empty) >= 2:
                return i
    return 0


def detect_csv_header_skip_rows(csv_path: Path, max_skip: int = 6) -> int:
    sample_rows: List[List[Any]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader):
            sample_rows.append(row)
            if idx >= max_skip + 1:
                break
    return detect_header_skip_rows(sample_rows, max_skip=max_skip)


def detect_xlsx_header_skip_rows(xlsx_path: Path, max_skip: int = 6) -> int:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    sample_rows: List[List[Any]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        sample_rows.append(list(row))
        if idx >= max_skip + 1:
            break
    wb.close()
    return detect_header_skip_rows(sample_rows, max_skip=max_skip)


def build_upload_view_sql(table_name: str, file_path: Path, skip_rows: int = 0) -> str:
    table_name = safe_identifier(table_name)
    ext = file_path.suffix.lower()
    file_literal = sql_quote_literal(str(file_path).replace("\\", "/"))

    if ext == ".csv":
        return (
            f"CREATE OR REPLACE VIEW {table_name} AS "
            f"SELECT * FROM read_csv("
            f"{file_literal}, auto_detect=true, sample_size=-1, header=true, skip={max(0, int(skip_rows))})"
        )
    if ext in {".parquet", ".pq"}:
        return f"CREATE OR REPLACE VIEW {table_name} AS SELECT * FROM read_parquet({file_literal})"
    raise HTTPException(status_code=400, detail="仅支持 CSV / XLSX / Parquet")


def convert_xlsx_to_csv(xlsx_path: Path, csv_path: Path, skip_rows: int = 0) -> None:
    # 逐行转换 xlsx，避免一次性载入内存
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx < max(0, int(skip_rows)):
                continue
            writer.writerow(["" if cell is None else cell for cell in row])
    wb.close()


def export_to_excel(query_sql: str, out_path: Path) -> None:
    cursor = con.execute(query_sql)
    columns = [desc[0] for desc in cursor.description]

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("result")
    ws.append(columns)

    while True:
        chunk = cursor.fetchmany(5000)
        if not chunk:
            break
        for row in chunk:
            ws.append(list(row))
    wb.save(str(out_path))


def export_to_csv_utf8_sig(query_sql: str, out_path: Path) -> None:
    """CSV 使用 utf-8-sig，Excel 双击打开中文不乱码"""
    cursor = con.execute(query_sql)
    columns = [desc[0] for desc in cursor.description]
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        while True:
            chunk = cursor.fetchmany(10000)
            if not chunk:
                break
            for row in chunk:
                writer.writerow(list(row))


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


def _runtime_debug_payload() -> Dict[str, Any]:
    db_ok = True
    db_error = ""
    try:
        with CON_LOCK:
            con.execute("SELECT 1").fetchone()
    except Exception as exc:
        db_ok = False
        db_error = str(exc)
    uploaded = {
        role: {
            "table_name": (meta or {}).get("table_name", ""),
            "path": (meta or {}).get("path", ""),
        }
        for role, meta in uploaded_meta.items()
    }
    return {
        "status": "ok" if db_ok else "degraded",
        "pid": os.getpid(),
        "runtime_seconds": int(max(0, time.time() - RUNTIME_STARTED_AT)),
        "runtime_port": os.getenv("APP_RUNTIME_PORT", ""),
        "runtime_log": os.getenv("APP_RUNTIME_LOG", ""),
        "db_file": str(DB_FILE.resolve()),
        "db_open_ok": db_ok,
        "db_error": db_error,
        "uploaded_roles": uploaded,
    }


@app.get("/debug/runtime")
def debug_runtime() -> Dict[str, Any]:
    return _runtime_debug_payload()


@app.get("/api/debug/runtime")
def api_debug_runtime() -> Dict[str, Any]:
    return _runtime_debug_payload()


@app.post("/upload")
async def upload_file(
    role: Literal["engineering", "traffic", "station"] = Form(...),
    file: UploadFile = File(...),
) -> Dict[str, Any]:
    original_filename = (file.filename or "").strip()
    table_name = get_table_name_by_role(role)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".parquet", ".pq"}:
        raise HTTPException(status_code=400, detail="仅支持 CSV / XLSX / Parquet")

    unique_name = f"{role}_{uuid.uuid4().hex}{suffix}"
    saved_path = DATA_DIR / unique_name

    # 分块写入磁盘，避免一次性读入内存
    with saved_path.open("wb") as f:
        while True:
            chunk = await file.read(8 * 1024 * 1024)
            if not chunk:
                break
            f.write(chunk)
    await file.close()

    view_source_path = saved_path
    skip_rows = 0
    if suffix == ".csv":
        try:
            skip_rows = detect_csv_header_skip_rows(saved_path, max_skip=6)
        except Exception:
            skip_rows = 0
    if suffix == ".xlsx":
        converted_csv = DATA_DIR / f"{role}_{uuid.uuid4().hex}.csv"
        try:
            skip_rows = detect_xlsx_header_skip_rows(saved_path, max_skip=6)
            convert_xlsx_to_csv(saved_path, converted_csv, skip_rows=skip_rows)
            view_source_path = converted_csv
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"XLSX 转换失败: {str(exc)}") from exc

    create_view_sql = build_upload_view_sql(table_name, view_source_path, skip_rows=skip_rows)
    try:
        con.execute(create_view_sql)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"注册 DuckDB VIEW 失败: {str(exc)}") from exc

    cols = get_columns(table_name)
    uploaded_meta[role] = {
        "table_name": table_name,
        "path": str(view_source_path),
    }
    return {
        "message": "上传成功并已注册为 DuckDB VIEW",
        "role": role,
        "table_name": table_name,
        "columns": cols,
        "original_filename": original_filename,
        "saved_path": str(saved_path.resolve()),
        "source_path": str(Path(view_source_path).resolve()),
        "header_skip_rows": skip_rows,
    }


@app.post("/preview")
def preview_data(payload: PreviewRequest) -> Dict[str, Any]:
    table_name = ensure_table_exists(payload.role)
    preview = execute_preview_query(
        table_name=table_name,
        limit=min(payload.limit, 100),
        column_filters=payload.column_filters,
    )
    source_path = (uploaded_meta.get(payload.role) or {}).get("path", "")
    return {"role": payload.role, "source_path": source_path, **preview}


@app.post("/upload_pasted")
def upload_pasted(payload: PasteUploadRequest) -> Dict[str, Any]:
    content = (payload.content or "").strip("\ufeff\r\n\t ")
    if not content:
        raise HTTPException(status_code=400, detail="粘贴内容不能为空")
    table_name = get_table_name_by_role(payload.role)

    # 优先按制表符解析，失败再回退逗号分隔
    sample = content.splitlines()[:8]
    skip_rows = 0
    rows_hint: List[List[Any]] = []
    for line in sample:
        parts = [c.strip() for c in line.split("\t")]
        rows_hint.append(parts)
    if rows_hint and max((len(r) for r in rows_hint), default=0) <= 1:
        rows_hint = [[c.strip() for c in line.split(",")] for line in sample]
    skip_rows = detect_header_skip_rows(rows_hint, max_skip=6)

    out_name = f"{payload.role}_{uuid.uuid4().hex}.csv"
    out_path = DATA_DIR / out_name
    try:
        with out_path.open("w", newline="", encoding="utf-8-sig") as f:
            f.write(content + ("\n" if not content.endswith("\n") else ""))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"写入粘贴内容失败: {str(exc)}") from exc

    create_view_sql = build_upload_view_sql(table_name, out_path, skip_rows=skip_rows)
    try:
        con.execute(create_view_sql)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"注册 DuckDB VIEW 失败: {str(exc)}") from exc

    cols = get_columns(table_name)
    uploaded_meta[payload.role] = {
        "table_name": table_name,
        "path": str(out_path),
    }
    return {
        "message": "粘贴导入成功并已注册为 DuckDB VIEW",
        "role": payload.role,
        "table_name": table_name,
        "columns": cols,
        "saved_path": str(out_path.resolve()),
        "source_path": str(out_path.resolve()),
        "header_skip_rows": skip_rows,
    }


@app.post("/filter")
def filter_data(payload: FilterRequest) -> Dict[str, Any]:
    table_name = ensure_table_exists(payload.role)
    where_clause, params = build_where_clause(table_name, payload.conditions)

    order_clause = ""
    if payload.sort_field:
        sort_column_map = {c.lower(): c for c in get_columns(table_name)}
        real_sort_field = sort_column_map.get(payload.sort_field.lower())
        if not real_sort_field:
            raise HTTPException(status_code=400, detail=f"排序字段不存在: {payload.sort_field}")
        order = "DESC" if payload.sort_order == "desc" else "ASC"
        order_clause = f" ORDER BY {sql_quote_ident(real_sort_field)} {order}"

    query = f"SELECT * FROM {safe_identifier(table_name)}{where_clause}{order_clause} LIMIT {payload.limit}"
    cursor = con.execute(query, params)
    columns = [d[0] for d in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]
    return {
        "role": payload.role,
        "columns": columns,
        "rows": data,
        "applied_conditions": [c.model_dump() for c in payload.conditions],
    }


@app.post("/preview_result")
def preview_result_data(payload: ResultPreviewRequest) -> Dict[str, Any]:
    if payload.result_key not in result_queries:
        raise HTTPException(status_code=404, detail="result_key 不存在，请先执行筛选")
    meta = result_queries[payload.result_key]
    view_or_sql = meta.get(payload.table_type)
    if not view_or_sql:
        raise HTTPException(status_code=400, detail=f"当前结果不包含 {payload.table_type} 数据")
    table_name = safe_identifier(view_or_sql)
    limit = min(payload.limit, 100)
    cursor = con.execute(f"SELECT * FROM {table_name} LIMIT {limit}")
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    total_row = con.execute(f"SELECT count(*) FROM {table_name}").fetchone()
    total = int(total_row[0]) if total_row and total_row[0] is not None else 0
    return {
        "table_type": payload.table_type,
        "columns": columns,
        "rows": [dict(zip(columns, row)) for row in rows],
        "total_row_count": total,
        "matching_row_count": total,
    }


def _traffic_base_after_preview(
    traffic_table: str,
    traffic_column_filters: List[PreviewColumnFilter],
) -> str:
    """话务表经预览列筛选后的子查询 SQL（无外层别名）"""
    prev_lit = build_preview_column_filters_literal(traffic_table, traffic_column_filters)
    return f"SELECT * FROM {safe_identifier(traffic_table)}{prev_lit}"


def _engineering_base_after_preview(
    engineering_table: str,
    engineering_column_filters: List[PreviewColumnFilter],
) -> str:
    """工参表经预览列筛选后的子查询 SQL（无外层别名）"""
    prev_lit = build_preview_column_filters_literal(engineering_table, engineering_column_filters)
    return f"SELECT * FROM {safe_identifier(engineering_table)}{prev_lit}"


def _station_base_after_preview(
    station_table: str,
    station_column_filters: List[PreviewColumnFilter],
) -> str:
    """选站表经预览列筛选后的子查询 SQL（无外层别名）"""
    prev_lit = build_preview_column_filters_literal(station_table, station_column_filters)
    return f"SELECT * FROM {safe_identifier(station_table)}{prev_lit}"


def _resolve_station_traffic_view_name(view_key: str) -> str:
    view_name = station_traffic_views.get((view_key or "").strip())
    if not view_name:
        raise HTTPException(status_code=404, detail="station_traffic_view_key 不存在，请先执行“预览选站话务”")
    safe = safe_identifier(view_name)
    try:
        con.execute(f"PRAGMA table_info({safe})").fetchall()
    except Exception as exc:
        raise HTTPException(status_code=404, detail="选站话务预览视图已失效，请重新点击“预览选站话务”") from exc
    return safe


@app.post("/preview_station_traffic")
def preview_station_traffic(payload: StationTrafficPreviewRequest) -> Dict[str, Any]:
    traffic_table = ensure_table_exists("traffic")
    station_table = ensure_table_exists("station")
    real_traffic_id = resolve_real_column(traffic_table, payload.traffic_id_field)
    real_station_id = resolve_real_column(station_table, payload.station_id_field)

    traffic_sub = _traffic_base_after_preview(traffic_table, payload.traffic_column_filters)
    station_sub = _station_base_after_preview(station_table, payload.station_column_filters)
    limit = min(payload.limit, 100)

    view_key = (payload.view_key or "").strip() or uuid.uuid4().hex
    view_name = f"tmp_station_traffic_{view_key}"
    query = f"""
        WITH station_ids AS (
            SELECT DISTINCT {sql_quote_ident(real_station_id)} AS join_id
            FROM ({station_sub}) s
        )
        SELECT t.*
        FROM ({traffic_sub}) t
        INNER JOIN station_ids ids
            ON t.{sql_quote_ident(real_traffic_id)} = ids.join_id
    """
    con.execute(f"CREATE OR REPLACE TEMP VIEW {safe_identifier(view_name)} AS {query}")
    station_traffic_views[view_key] = view_name
    preview = execute_preview_query(
        table_name=view_name,
        limit=limit,
        column_filters=payload.column_filters,
    )
    traffic_source = (uploaded_meta.get("traffic") or {}).get("path", "")
    station_source = (uploaded_meta.get("station") or {}).get("path", "")
    return {
        "role": "traffic",
        **preview,
        "traffic_source_path": traffic_source,
        "station_source_path": station_source,
        "view_key": view_key,
        "view_name": view_name,
        "joined_by": {
            "traffic_id_field": real_traffic_id,
            "station_id_field": real_station_id,
        },
    }


@app.post("/preview_station_traffic_column_distinct")
def preview_station_traffic_column_distinct(payload: StationTrafficDistinctRequest) -> Dict[str, Any]:
    view_name = _resolve_station_traffic_view_name(payload.view_key)
    real_field = resolve_real_column(view_name, payload.field)
    where_clause, params = build_preview_column_filters_where(view_name, payload.column_filters)
    try:
        rows = con.execute(
            f"""
            SELECT DISTINCT {sql_quote_ident(real_field)} AS v
            FROM {view_name}
            {where_clause}
            ORDER BY CAST(v AS VARCHAR)
            LIMIT {payload.max_values}
            """,
            params,
        ).fetchall()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取去重值失败: {str(exc)}") from exc
    values = [r[0] for r in rows]
    return {
        "field": real_field,
        "values": values,
        "truncated": len(values) >= payload.max_values,
    }


@app.post("/preview_station_traffic_distinct_count")
def preview_station_traffic_distinct_count(payload: StationTrafficDistinctRequest) -> Dict[str, Any]:
    view_name = _resolve_station_traffic_view_name(payload.view_key)
    real_field = resolve_real_column(view_name, payload.field)
    where_clause, params = build_preview_column_filters_where(view_name, payload.column_filters)
    try:
        row = con.execute(
            f"""
            SELECT count(DISTINCT {sql_quote_ident(real_field)})::BIGINT AS cnt
            FROM {view_name}
            {where_clause}
            """,
            params,
        ).fetchone()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"统计去重数失败: {str(exc)}") from exc
    return {
        "field": real_field,
        "distinct_count": int(row[0]) if row and row[0] is not None else 0,
    }


@app.post("/station_distinct_count")
def station_distinct_count(payload: StationDistinctCountRequest) -> Dict[str, Any]:
    station_table = ensure_table_exists("station")
    real_field = resolve_real_column(station_table, payload.field)
    try:
        row = con.execute(
            f"""
            SELECT count(DISTINCT {sql_quote_ident(real_field)})::BIGINT AS cnt
            FROM {safe_identifier(station_table)}
            """
        ).fetchone()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"统计选站去重数失败: {str(exc)}") from exc
    return {
        "field": real_field,
        "distinct_count": int(row[0]) if row and row[0] is not None else 0,
    }


@app.post("/condition_chart_counts")
def condition_chart_counts(payload: ConditionChartCountRequest) -> Dict[str, Any]:
    """统计条件组与单条件在“当前表格基准数据”上的筛选前后数量。"""
    traffic_table = ensure_table_exists("traffic")
    engineering_table = ensure_table_exists("engineering") if "engineering" in uploaded_meta else None
    station_table = ensure_table_exists("station") if "station" in uploaded_meta else None

    if payload.condition_role == "traffic":
        cond_table = traffic_table
        cond_inner = _traffic_base_after_preview(traffic_table, payload.traffic_column_filters)
    elif payload.condition_role == "engineering":
        if not engineering_table:
            raise HTTPException(status_code=400, detail="工参数据未上传，无法统计条件组图")
        cond_table = engineering_table
        cond_inner = _engineering_base_after_preview(engineering_table, payload.engineering_column_filters)
    else:
        if payload.station_traffic_view_key:
            cond_table = _resolve_station_traffic_view_name(payload.station_traffic_view_key)
            cond_inner = f"SELECT * FROM {cond_table}"
        else:
            if not station_table:
                raise HTTPException(status_code=400, detail="选站数据未上传，无法统计条件组图")
            cond_table = station_table
            cond_inner = _station_base_after_preview(station_table, payload.station_column_filters)

    base_row = con.execute(f"SELECT count(*) FROM ({cond_inner}) base").fetchone()
    base_count = int(base_row[0]) if base_row and base_row[0] is not None else 0

    group_where = build_where_clause_literal(cond_table, payload.conditions)
    group_row = con.execute(f"SELECT count(*) FROM ({cond_inner}) base{group_where}").fetchone()
    group_filtered_count = int(group_row[0]) if group_row and group_row[0] is not None else 0

    one_condition_counts: List[Dict[str, Any]] = []
    for idx, cond in enumerate(payload.conditions):
        one_where = build_where_clause_literal(cond_table, [cond])
        one_row = con.execute(f"SELECT count(*) FROM ({cond_inner}) base{one_where}").fetchone()
        one_filtered = int(one_row[0]) if one_row and one_row[0] is not None else 0
        one_condition_counts.append(
            {
                "index": idx,
                "field": cond.field,
                "operator": cond.operator,
                "value": cond.value,
                "filtered_count": one_filtered,
            }
        )

    return {
        "condition_role": payload.condition_role,
        "base_count": base_count,
        "group_filtered_count": group_filtered_count,
        "one_condition_counts": one_condition_counts,
    }


@app.post("/nl_filter")
def nl_filter(payload: NlFilterRequest) -> Dict[str, Any]:
    traffic_table = ensure_table_exists(payload.traffic_role)
    real_traffic_id = resolve_real_column(traffic_table, payload.traffic_id_field)
    traffic_inner = _traffic_base_after_preview(traffic_table, payload.traffic_column_filters)
    engineering_table = ensure_table_exists(payload.engineering_role) if payload.engineering_role in uploaded_meta else None
    station_table = ensure_table_exists(payload.station_role) if payload.station_role in uploaded_meta else None
    station_traffic_view_name = (
        _resolve_station_traffic_view_name(payload.station_traffic_view_key)
        if payload.station_traffic_view_key
        else None
    )

    engineering_inner = (
        _engineering_base_after_preview(engineering_table, payload.engineering_column_filters)
        if engineering_table
        else None
    )
    station_inner = (
        _station_base_after_preview(station_table, payload.station_column_filters)
        if station_table
        else None
    )
    real_engineering_id = (
        resolve_real_column(engineering_table, payload.engineering_id_field) if engineering_table else None
    )
    real_station_id = resolve_real_column(station_table, payload.station_id_field) if station_table else None

    groups = payload.condition_groups or []
    if not groups:
        if not payload.text or not payload.text.strip():
            raise HTTPException(status_code=400, detail="请提供自然语言 text 或 condition_groups")
        groups = parse_nl_conditions(payload.text)
    group_results: List[Dict[str, Any]] = []

    for idx, conds in enumerate(groups, start=1):
        # 可切换条件作用于话务 / 工参 / 选站，统一关联回查话务
        if payload.condition_role == "traffic":
            cond_table = traffic_table
            cond_inner = traffic_inner
            cond_id = real_traffic_id
        elif payload.condition_role == "engineering":
            if not engineering_table or not engineering_inner or not real_engineering_id:
                raise HTTPException(status_code=400, detail="工参数据未上传，无法按工参条件关联筛选")
            cond_table = engineering_table
            cond_inner = engineering_inner
            cond_id = real_engineering_id
        else:
            if station_traffic_view_name:
                cond_table = station_traffic_view_name
                cond_inner = f"SELECT * FROM {station_traffic_view_name}"
                cond_id = resolve_real_column(station_traffic_view_name, payload.traffic_id_field)
            else:
                if not station_table or not station_inner or not real_station_id:
                    raise HTTPException(status_code=400, detail="选站数据未上传，无法按选站条件关联筛选")
                cond_table = station_table
                cond_inner = station_inner
                cond_id = real_station_id

        where_clause_literal = build_where_clause_literal(cond_table, conds)
        if where_clause_literal:
            cond_query_literal = f"SELECT * FROM ({cond_inner}) AS cond_base{where_clause_literal}"
        else:
            cond_query_literal = cond_inner

        result_key = f"nl_{uuid.uuid4().hex}"
        group_name = None
        if payload.group_names and len(payload.group_names) >= idx:
            group_name = (payload.group_names[idx - 1] or "").strip()
        group_name = group_name or f"结果组{idx}"

        # 关键：把带参数的查询先落成临时 VIEW，避免导出时丢失参数
        traffic_view = f"tmp_{result_key}_traffic"
        traffic_ids_view = f"tmp_{result_key}_traffic_ids"
        eng_view = f"tmp_{result_key}_engineering"
        station_view = f"tmp_{result_key}_station"
        if payload.condition_role == "traffic":
            con.execute(f"CREATE OR REPLACE TEMP VIEW {safe_identifier(traffic_view)} AS {cond_query_literal}")
            con.execute(
                f"CREATE OR REPLACE TEMP VIEW {safe_identifier(traffic_ids_view)} AS "
                f"SELECT DISTINCT {sql_quote_ident(real_traffic_id)} AS join_id FROM {safe_identifier(traffic_view)}"
            )
        else:
            con.execute(
                f"CREATE OR REPLACE TEMP VIEW {safe_identifier(traffic_ids_view)} AS "
                f"SELECT DISTINCT {sql_quote_ident(cond_id)} AS join_id FROM ({cond_query_literal}) src"
            )
            con.execute(
                f"CREATE OR REPLACE TEMP VIEW {safe_identifier(traffic_view)} AS "
                f"SELECT t.* FROM ({traffic_inner}) t "
                f"INNER JOIN {safe_identifier(traffic_ids_view)} ids ON "
                f"t.{sql_quote_ident(real_traffic_id)} = ids.join_id"
            )
            if payload.condition_role == "engineering":
                con.execute(f"CREATE OR REPLACE TEMP VIEW {safe_identifier(eng_view)} AS {cond_query_literal}")
            elif payload.condition_role == "station":
                con.execute(f"CREATE OR REPLACE TEMP VIEW {safe_identifier(station_view)} AS {cond_query_literal}")

        if payload.condition_role == "traffic" and engineering_inner and real_engineering_id:
            con.execute(
                f"CREATE OR REPLACE TEMP VIEW {safe_identifier(eng_view)} AS "
                f"SELECT e.* FROM ({engineering_inner}) e "
                f"INNER JOIN {safe_identifier(traffic_ids_view)} ids ON "
                f"e.{sql_quote_ident(real_engineering_id)} = ids.join_id"
            )
        elif payload.condition_role != "engineering" and engineering_inner and real_engineering_id:
            con.execute(
                f"CREATE OR REPLACE TEMP VIEW {safe_identifier(eng_view)} AS "
                f"SELECT e.* FROM ({engineering_inner}) e "
                f"INNER JOIN {safe_identifier(traffic_ids_view)} ids ON "
                f"e.{sql_quote_ident(real_engineering_id)} = ids.join_id"
            )

        result_queries[result_key] = {
            "traffic": traffic_view,
            "engineering": eng_view if engineering_inner else "",
            "station": station_view if payload.condition_role == "station" and not station_traffic_view_name else "",
            "group_name": group_name,
        }

        traffic_row_count = int(
            con.execute(f"SELECT count(*) FROM {safe_identifier(traffic_view)}").fetchone()[0]
        )
        traffic_distinct_cell_count = int(
            con.execute(
                f"SELECT count(DISTINCT {sql_quote_ident(real_traffic_id)}) "
                f"FROM {safe_identifier(traffic_view)}"
            ).fetchone()[0]
        )
        engineering_row_count = (
            int(con.execute(f"SELECT count(*) FROM {safe_identifier(eng_view)}").fetchone()[0])
            if engineering_inner
            else 0
        )
        station_row_count = (
            int(con.execute(f"SELECT count(*) FROM {safe_identifier(station_view)}").fetchone()[0])
            if payload.condition_role == "station" and not station_traffic_view_name
            else 0
        )

        group_results.append(
            {
                "group_index": idx,
                "result_key": result_key,
                "group_name": group_name,
                "conditions": [c.model_dump() for c in conds],
                "condition_role": payload.condition_role,
                "stats": {
                    "traffic_row_count": traffic_row_count,
                    "traffic_distinct_cell_count": traffic_distinct_cell_count,
                    "engineering_row_count": engineering_row_count,
                    "station_row_count": station_row_count,
                },
            }
        )

    return {"groups": group_results}


@app.post("/nl_filter_parse")
def nl_filter_parse(payload: NlParseRequest) -> Dict[str, Any]:
    parsed_groups = parse_nl_conditions(payload.text)
    traffic_table = ensure_table_exists(payload.traffic_role)
    traffic_cols = get_columns(traffic_table)
    traffic_col_map = {c.lower(): c for c in traffic_cols}

    normalized_groups: List[List[Dict[str, Any]]] = []
    unmatched_fields: List[str] = []
    for group in parsed_groups:
        current_group: List[Dict[str, Any]] = []
        for cond in group:
            mapped_field = traffic_col_map.get(cond.field.lower())
            if not mapped_field:
                unmatched_fields.append(cond.field)
                mapped_field = cond.field
            current_group.append(
                {
                    "field": mapped_field,
                    "operator": cond.operator,
                    "value": cond.value,
                }
            )
        normalized_groups.append(current_group)

    return {
        "groups": normalized_groups,
        "available_fields": traffic_cols,
        "unmatched_fields": sorted(list(set(unmatched_fields))),
    }


@app.post("/field_max")
def field_max(payload: FieldMaxRequest) -> Dict[str, Any]:
    table_name = ensure_table_exists(payload.role)
    real_field = resolve_real_column(table_name, payload.field)
    col_type = (get_column_types(table_name).get(real_field) or "").upper()
    try:
        if any(t in col_type for t in ["VARCHAR", "TEXT", "STRING"]):
            # 忽略 NIL/空值；将可转数值的字符串参与最大值计算
            row = con.execute(
                f"""
                SELECT max(try_cast(nullif(trim({sql_quote_ident(real_field)}), 'NIL') AS DOUBLE)) AS max_value
                FROM {safe_identifier(table_name)}
                """
            ).fetchone()
        else:
            row = con.execute(
                f"SELECT max({sql_quote_ident(real_field)}) AS max_value FROM {safe_identifier(table_name)}"
            ).fetchone()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"计算最大值失败: {str(exc)}") from exc

    return {
        "role": payload.role,
        "field": real_field,
        "max_value": row[0] if row else None,
    }


@app.post("/column_distinct")
def column_distinct(payload: ColumnDistinctRequest) -> Dict[str, Any]:
    """列头筛选：返回该列在全表中的去重值（用于下拉选项，不限于预览 100 行）"""
    table_name = ensure_table_exists(payload.role)
    real_field = resolve_real_column(table_name, payload.field)
    try:
        rows = con.execute(
            f"""
            SELECT DISTINCT {sql_quote_ident(real_field)} AS v
            FROM {safe_identifier(table_name)}
            ORDER BY CAST(v AS VARCHAR)
            LIMIT {payload.max_values}
            """
        ).fetchall()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取去重值失败: {str(exc)}") from exc
    values = [r[0] for r in rows]
    return {
        "field": real_field,
        "values": values,
        "truncated": len(values) >= payload.max_values,
    }


@app.post("/export_filtered_preview")
def export_filtered_preview(payload: ExportFilteredRequest) -> Dict[str, Any]:
    """工参/话务：按列筛选导出全表行；无筛选条件时导出整张表"""
    if payload.file_format not in ALLOWED_EXPORT_FORMATS:
        raise HTTPException(status_code=400, detail="导出格式仅支持 csv 或 excel")
    table_name = ensure_table_exists(payload.role)
    if payload.column_filters:
        where_lit = build_preview_column_filters_literal(table_name, payload.column_filters)
        base_query = f"SELECT * FROM {safe_identifier(table_name)}{where_lit}"
    else:
        base_query = f"SELECT * FROM {safe_identifier(table_name)}"

    ext = "xlsx" if payload.file_format == "excel" else payload.file_format
    role_cn_map = {"engineering": "工参", "traffic": "话务", "station": "选站"}
    role_cn = role_cn_map.get(payload.role, payload.role)
    out_name = f"预览筛选_{role_cn}_{uuid.uuid4().hex[:10]}.{ext}"
    out_path = EXPORT_DIR / out_name

    try:
        if payload.file_format == "csv":
            export_to_csv_utf8_sig(base_query, out_path)
        else:
            export_to_excel(base_query, out_path)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(exc)}") from exc

    return {
        "message": "导出成功",
        "file_name": out_name,
        "file_path": str(out_path),
    }


@app.post("/export")
def export_result(payload: ExportRequest) -> Dict[str, Any]:
    if payload.file_format not in ALLOWED_EXPORT_FORMATS:
        raise HTTPException(status_code=400, detail="导出格式仅支持 csv 或 excel")
    if payload.result_key not in result_queries:
        raise HTTPException(status_code=404, detail="result_key 不存在，请先执行筛选")

    meta = result_queries[payload.result_key]
    view_or_sql = meta.get(payload.table_type)
    if not view_or_sql or not isinstance(view_or_sql, str):
        raise HTTPException(status_code=400, detail="table_type 仅支持 traffic 或 engineering")

    ext = "xlsx" if payload.file_format == "excel" else payload.file_format
    group_name = safe_filename(meta.get("group_name", payload.result_key))
    type_cn = "话务" if payload.table_type == "traffic" else "工参"
    out_name = f"{group_name}_{type_cn}.{ext}"
    out_path = EXPORT_DIR / out_name

    try:
        # 关联结果视图名由服务端生成，仅含字母数字下划线，用 safe_identifier 最稳
        query_sql = f"SELECT * FROM {safe_identifier(view_or_sql)}"
        if payload.file_format == "csv":
            export_to_csv_utf8_sig(query_sql, out_path)
        else:
            export_to_excel(query_sql, out_path)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(exc)}") from exc

    return {
        "message": "导出成功",
        "file_name": out_name,
        "file_path": str(out_path),
    }


@app.post("/export_nl_batch")
def export_nl_batch(payload: ExportNlBatchRequest) -> Dict[str, Any]:
    """一键导出所有条件组：Excel 为每条件组一个工作表；CSV 为单行合并并带「条件组名称」列"""
    if not payload.groups:
        raise HTTPException(status_code=400, detail="groups 不能为空")
    if payload.file_format not in ALLOWED_EXPORT_FORMATS:
        raise HTTPException(status_code=400, detail="导出格式仅支持 csv 或 excel")

    table_type = payload.table_type
    for item in payload.groups:
        if item.result_key not in result_queries:
            raise HTTPException(status_code=404, detail=f"result_key 不存在: {item.result_key}")
        meta = result_queries[item.result_key]
        if not meta.get(table_type):
            raise HTTPException(status_code=400, detail=f"缺少 {table_type} 结果: {item.result_key}")

    type_cn = "话务" if table_type == "traffic" else "工参"
    batch_id = uuid.uuid4().hex[:10]

    if payload.file_format == "excel":
        out_name = f"全部条件组_{type_cn}_{batch_id}.xlsx"
        out_path = EXPORT_DIR / out_name
        wb = Workbook()
        wb.remove(wb.active)

        for idx, item in enumerate(payload.groups, start=1):
            meta = result_queries[item.result_key]
            view_or_sql = meta[table_type]
            title = safe_excel_sheet_title(
                f"{idx}_{item.group_name or meta.get('group_name', '组')}"
            )

            ws = wb.create_sheet(title=title)
            cursor = con.execute(f"SELECT * FROM {safe_identifier(view_or_sql)}")
            columns = [desc[0] for desc in cursor.description]
            ws.append(columns)
            while True:
                chunk = cursor.fetchmany(5000)
                if not chunk:
                    break
                for row in chunk:
                    ws.append(list(row))

        wb.save(str(out_path))
    else:
        out_name = f"全部条件组_{type_cn}_{batch_id}.csv"
        out_path = EXPORT_DIR / out_name
        union_parts: List[str] = []
        for item in payload.groups:
            meta = result_queries[item.result_key]
            view_or_sql = meta[table_type]
            gname = safe_filename(item.group_name or meta.get("group_name", "组"))
            union_parts.append(
                f'SELECT {sql_literal(gname)} AS "条件组名称", t.* FROM '
                f"{safe_identifier(view_or_sql)} t"
            )
        combined = " UNION ALL ".join(union_parts)
        export_to_csv_utf8_sig(combined, out_path)

    return {
        "message": "导出成功",
        "file_name": out_name,
        "file_path": str(out_path),
    }


@app.post("/engineering_chart")
def engineering_chart(payload: EngineeringChartRequest) -> Dict[str, Any]:
    """工参列分布（分类 TopN）或数值阈值占比（在预览列筛选子集上统计）"""
    engineering_table = ensure_table_exists("engineering")
    real_field = resolve_real_column(engineering_table, payload.field)
    col_sql = sql_quote_ident(real_field)
    base_lit = build_preview_column_filters_literal(engineering_table, payload.column_filters)
    base_sub = f"SELECT * FROM {safe_identifier(engineering_table)}{base_lit}"

    if payload.mode == "threshold":
        if payload.operator is None or payload.threshold_value is None:
            raise HTTPException(status_code=400, detail="threshold 模式需同时提供 operator 与 threshold_value")
        op = payload.operator
        val = payload.threshold_value
        # 中文注释：阈值比较统一按数值 try_cast，与条件组筛选一致
        field_ref = f"t.{col_sql}"
        q = f"""
            SELECT
              count(*) AS total,
              coalesce(
                sum(CASE WHEN try_cast({field_ref} AS DOUBLE) {op} ? THEN 1 ELSE 0 END),
                0
              )::BIGINT AS match_cnt
            FROM ({base_sub}) t
        """
        row = con.execute(q, [val]).fetchone()
        total = int(row[0]) if row and row[0] is not None else 0
        match_cnt = int(row[1]) if row and row[1] is not None else 0
        prop = (match_cnt / total) if total else 0.0
        return {
            "mode": "threshold",
            "field": real_field,
            "total_rows": total,
            "match_count": match_cnt,
            "proportion": prop,
            "operator": op,
            "threshold_value": val,
        }

    # top_values：按字符串化后的取值聚合
    try:
        field_ref = f"t.{col_sql}"
        rows = con.execute(
            f"""
            SELECT CAST({field_ref} AS VARCHAR) AS bucket, count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            GROUP BY bucket
            ORDER BY cnt DESC
            LIMIT {int(payload.top_n)}
            """
        ).fetchall()
        total_row = con.execute(f"SELECT count(*) FROM ({base_sub}) t").fetchone()
        total = int(total_row[0]) if total_row and total_row[0] is not None else 0
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"统计失败: {str(exc)}") from exc

    items: List[Dict[str, Any]] = []
    for r in rows:
        bucket = r[0] if r[0] is not None else ""
        label = "(空)" if bucket == "" else str(bucket)
        cnt = int(r[1])
        pct = (100.0 * cnt / total) if total else 0.0
        items.append({"label": label, "count": cnt, "percent": pct})

    return {
        "mode": "top_values",
        "field": real_field,
        "total_rows": total,
        "items": items,
        "truncated_to_top_n": len(items) >= payload.top_n,
    }


@app.post("/column_distribution")
def column_distribution(payload: ColumnDistributionRequest) -> Dict[str, Any]:
    """表格列点击即看：自动识别类别/数值；返回分类分布或数值分布"""
    table_name = ensure_table_exists(payload.role)
    real_field = resolve_real_column(table_name, payload.field)
    col_sql = sql_quote_ident(real_field)
    col_types = get_column_types(table_name)
    col_type = (col_types.get(real_field) or "").upper()
    is_numeric_declared = any(
        t in col_type
        for t in [
            "INT",
            "DOUBLE",
            "FLOAT",
            "DECIMAL",
            "NUMERIC",
            "HUGEINT",
            "UBIGINT",
            "SMALLINT",
            "TINYINT",
            "BIGINT",
        ]
    )

    base_lit = build_preview_column_filters_literal(table_name, payload.column_filters)
    base_sub = f"SELECT * FROM {safe_identifier(table_name)}{base_lit}"
    total_row = con.execute(f"SELECT count(*) FROM ({base_sub}) t").fetchone()
    total = int(total_row[0]) if total_row and total_row[0] is not None else 0

    # 自动识别：声明为数值类型，或字符串列中可转数值占比足够高
    is_numeric = is_numeric_declared
    if not is_numeric:
        probe = con.execute(
            f"""
            SELECT
              count(*)::BIGINT AS total_cnt,
              count(try_cast(t.{col_sql} AS DOUBLE))::BIGINT AS numeric_cnt
            FROM ({base_sub}) t
            """
        ).fetchone()
        total_cnt = int(probe[0]) if probe and probe[0] is not None else 0
        numeric_cnt = int(probe[1]) if probe and probe[1] is not None else 0
        is_numeric = total_cnt > 0 and (numeric_cnt / total_cnt) >= 0.8

    if (not is_numeric) or payload.mode == "auto":
        if not is_numeric:
            rows = con.execute(
                f"""
                SELECT CAST(t.{col_sql} AS VARCHAR) AS bucket, count(*)::BIGINT AS cnt
                FROM ({base_sub}) t
                GROUP BY bucket
                ORDER BY cnt DESC
                LIMIT {int(payload.top_n)}
                """
            ).fetchall()
            items: List[Dict[str, Any]] = []
            for r in rows:
                bucket = r[0] if r[0] is not None else ""
                label = "(空)" if bucket == "" else str(bucket)
                cnt = int(r[1])
                pct = (100.0 * cnt / total) if total else 0.0
                items.append({"label": label, "count": cnt, "percent": pct})
            return {
                "role": payload.role,
                "field": real_field,
                "detected_type": "categorical",
                "mode": "top_values",
                "total_rows": total,
                "items": items,
            }
        # auto 且识别为数值时，默认走直方图
        mode = "histogram"
    else:
        mode = payload.mode

    if mode == "threshold_3bins":
        if payload.x1 is None and payload.x2 is None:
            raise HTTPException(status_code=400, detail="阈值模式至少需要提供 x1 或 x2")
        x1 = float(payload.x1 if payload.x1 is not None else payload.x2)
        x2 = float(payload.x2 if payload.x2 is not None else payload.x1)
        if x1 > x2:
            x1, x2 = x2, x1
        two_bins = abs(x1 - x2) < 1e-12
        row = con.execute(
            f"""
            SELECT
              coalesce(sum(CASE WHEN try_cast(t.{col_sql} AS DOUBLE) < ? THEN 1 ELSE 0 END),0)::BIGINT AS c1,
              coalesce(sum(CASE WHEN try_cast(t.{col_sql} AS DOUBLE) >= ? AND try_cast(t.{col_sql} AS DOUBLE) <= ? THEN 1 ELSE 0 END),0)::BIGINT AS c2,
              coalesce(sum(CASE WHEN try_cast(t.{col_sql} AS DOUBLE) > ? THEN 1 ELSE 0 END),0)::BIGINT AS c3
            FROM ({base_sub}) t
            """,
            [x1, x1, x2, x2],
        ).fetchone()
        c1 = int(row[0]) if row and row[0] is not None else 0
        c2 = int(row[1]) if row and row[1] is not None else 0
        c3 = int(row[2]) if row and row[2] is not None else 0
        if two_bins:
            items = [
                {"label": f"(< {x1})", "count": c1},
                {"label": f"(>= {x1})", "count": c2 + c3},
            ]
        else:
            items = [
                {"label": f"(< {x1})", "count": c1},
                {"label": f"[{x1}, {x2}]", "count": c2},
                {"label": f"(> {x2})", "count": c3},
            ]
        return {
            "role": payload.role,
            "field": real_field,
            "detected_type": "numeric",
            "mode": "threshold_3bins",
            "total_rows": total,
            "x1": x1,
            "x2": x2,
            "items": items,
        }

    # histogram
    bins = int(payload.bins)
    mm = con.execute(
        f"""
        SELECT
          min(try_cast(t.{col_sql} AS DOUBLE)) AS min_v,
          max(try_cast(t.{col_sql} AS DOUBLE)) AS max_v
        FROM ({base_sub}) t
        """
    ).fetchone()
    min_v = float(mm[0]) if mm and mm[0] is not None else None
    max_v = float(mm[1]) if mm and mm[1] is not None else None
    if min_v is None or max_v is None:
        return {
            "role": payload.role,
            "field": real_field,
            "detected_type": "numeric",
            "mode": "histogram",
            "total_rows": total,
            "items": [],
        }

    if max_v == min_v:
        rows = con.execute(
            f"""
            SELECT count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            WHERE try_cast(t.{col_sql} AS DOUBLE) IS NOT NULL
            """
        ).fetchone()
        only_cnt = int(rows[0]) if rows and rows[0] is not None else 0
        return {
            "role": payload.role,
            "field": real_field,
            "detected_type": "numeric",
            "mode": "histogram",
            "total_rows": total,
            "min": min_v,
            "max": max_v,
            "items": [{"label": f"[{min_v}, {max_v}]", "count": only_cnt}],
        }

    width = (max_v - min_v) / bins
    rows = con.execute(
        f"""
        WITH s AS (
          SELECT try_cast(t.{col_sql} AS DOUBLE) AS v
          FROM ({base_sub}) t
          WHERE try_cast(t.{col_sql} AS DOUBLE) IS NOT NULL
        )
        SELECT
          LEAST(FLOOR((v - ?) / ?), ?)::INT AS bucket,
          count(*)::BIGINT AS cnt
        FROM s
        GROUP BY bucket
        ORDER BY bucket
        """,
        [min_v, width, bins - 1],
    ).fetchall()
    items: List[Dict[str, Any]] = []
    for bucket, cnt in rows:
        b = int(bucket)
        left = min_v + b * width
        right = max_v if b == bins - 1 else (left + width)
        items.append(
            {
                "bucket": b,
                "left": left,
                "right": right,
                "label": f"[{left:.3f}, {right:.3f}{']' if b == bins - 1 else ')'}",
                "count": int(cnt),
            }
        )
    return {
        "role": payload.role,
        "field": real_field,
        "detected_type": "numeric",
        "mode": "histogram",
        "total_rows": total,
        "min": min_v,
        "max": max_v,
        "bins": bins,
        "items": items,
    }


@app.post("/distribution_compare")
def distribution_compare(payload: CompareDistributionRequest) -> Dict[str, Any]:
    """返回全量与条件过滤后两组分布，用于双层柱状对比"""
    table_name = resolve_table_name_input(payload.table_name)
    real_field = resolve_real_column(table_name, payload.field)
    col_sql = sql_quote_ident(real_field)
    mode = payload.mode

    # 自动识别：数值列走 histogram，其他走 top_values
    if mode == "auto":
        col_type = (get_column_types(table_name).get(real_field) or "").upper()
        is_numeric = any(
            t in col_type
            for t in ["INT", "DOUBLE", "FLOAT", "DECIMAL", "NUMERIC", "BIGINT", "SMALLINT", "TINYINT"]
        )
        mode = "histogram" if is_numeric else "top_values"

    base_sub = f"SELECT * FROM {safe_identifier(table_name)}"
    cond_lit = build_where_clause_literal(table_name, payload.condition_list)
    filtered_sub = f"SELECT * FROM {safe_identifier(table_name)}{cond_lit}"

    if mode == "top_values":
        labels_rows = con.execute(
            f"""
            SELECT CAST(t.{col_sql} AS VARCHAR) AS label, count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            GROUP BY label
            ORDER BY cnt DESC
            LIMIT {int(payload.top_n)}
            """
        ).fetchall()
        labels = [(r[0] if r[0] is not None else "") for r in labels_rows]
        if not labels:
            return {
                "mode": "top_values",
                "field": real_field,
                "detected_type": "categorical",
                "items": [],
            }
        in_clause = ",".join(sql_literal(v) for v in labels)
        full_rows = con.execute(
            f"""
            SELECT CAST(t.{col_sql} AS VARCHAR) AS label, count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            WHERE CAST(t.{col_sql} AS VARCHAR) IN ({in_clause})
            GROUP BY label
            """
        ).fetchall()
        filt_rows = con.execute(
            f"""
            SELECT CAST(t.{col_sql} AS VARCHAR) AS label, count(*)::BIGINT AS cnt
            FROM ({filtered_sub}) t
            WHERE CAST(t.{col_sql} AS VARCHAR) IN ({in_clause})
            GROUP BY label
            """
        ).fetchall()
        full_map = {("" if r[0] is None else str(r[0])): int(r[1]) for r in full_rows}
        filt_map = {("" if r[0] is None else str(r[0])): int(r[1]) for r in filt_rows}
        items = []
        for lb in labels:
            k = "" if lb is None else str(lb)
            shown = "(空)" if k == "" else k
            items.append(
                {
                    "label": shown,
                    "full_count": full_map.get(k, 0),
                    "filtered_count": filt_map.get(k, 0),
                }
            )
        return {
            "mode": "top_values",
            "field": real_field,
            "detected_type": "categorical",
            "items": items,
        }

    # histogram compare
    bins = int(payload.bins)
    mm = con.execute(
        f"""
        SELECT
          min(try_cast(t.{col_sql} AS DOUBLE)) AS min_v,
          max(try_cast(t.{col_sql} AS DOUBLE)) AS max_v
        FROM ({base_sub}) t
        """
    ).fetchone()
    min_v = float(mm[0]) if mm and mm[0] is not None else None
    max_v = float(mm[1]) if mm and mm[1] is not None else None
    if min_v is None or max_v is None:
        return {"mode": "histogram", "field": real_field, "detected_type": "numeric", "items": []}
    if max_v == min_v:
        full_cnt = int(
            con.execute(
                f"SELECT count(*) FROM ({base_sub}) t WHERE try_cast(t.{col_sql} AS DOUBLE) IS NOT NULL"
            ).fetchone()[0]
        )
        filt_cnt = int(
            con.execute(
                f"SELECT count(*) FROM ({filtered_sub}) t WHERE try_cast(t.{col_sql} AS DOUBLE) IS NOT NULL"
            ).fetchone()[0]
        )
        return {
            "mode": "histogram",
            "field": real_field,
            "detected_type": "numeric",
            "items": [{"label": f"[{min_v}, {max_v}]", "full_count": full_cnt, "filtered_count": filt_cnt}],
        }
    width = (max_v - min_v) / bins

    def bucket_counts(sub_sql: str) -> Dict[int, int]:
        rs = con.execute(
            f"""
            WITH s AS (
              SELECT try_cast(t.{col_sql} AS DOUBLE) AS v
              FROM ({sub_sql}) t
              WHERE try_cast(t.{col_sql} AS DOUBLE) IS NOT NULL
            )
            SELECT LEAST(FLOOR((v - ?) / ?), ?)::INT AS bucket, count(*)::BIGINT AS cnt
            FROM s
            GROUP BY bucket
            """,
            [min_v, width, bins - 1],
        ).fetchall()
        return {int(r[0]): int(r[1]) for r in rs}

    full_map = bucket_counts(base_sub)
    filt_map = bucket_counts(filtered_sub)
    items = []
    for b in range(bins):
        left = min_v + b * width
        right = max_v if b == bins - 1 else (left + width)
        items.append(
            {
                "label": f"[{left:.3f}, {right:.3f}{']' if b == bins - 1 else ')'}",
                "full_count": full_map.get(b, 0),
                "filtered_count": filt_map.get(b, 0),
            }
        )
    return {"mode": "histogram", "field": real_field, "detected_type": "numeric", "items": items}


def _distribution_core(payload: UnifiedDistributionRequest) -> Dict[str, Any]:
    table_name = resolve_table_name_input(payload.table_name)
    real_field = resolve_real_column(table_name, payload.column)
    col_sql = sql_quote_ident(real_field)
    mode = payload.mode

    base_sub = f"SELECT * FROM {safe_identifier(table_name)}"
    preview_lit = build_preview_column_filters_literal(table_name, payload.preview_filters)
    filtered_preview_sub = f"SELECT * FROM {safe_identifier(table_name)}{preview_lit}"
    col_name_set = {c.lower() for c in get_columns(table_name)}
    safe_conds = [c for c in payload.conditions if c.field.lower() in col_name_set]
    cond_lit = build_where_clause_literal(table_name, safe_conds)
    if cond_lit:
        filtered_sub = f"SELECT * FROM ({filtered_preview_sub}) t{cond_lit}"
    else:
        filtered_sub = filtered_preview_sub
    compare_effective = bool(payload.compare_with_base and (safe_conds or payload.preview_filters))

    junk_w = _junk_where_varchar(col_sql)
    fin_w = _finite_double_where(col_sql)

    if mode == "auto":
        col_type = (get_column_types(table_name).get(real_field) or "").upper()
        pragma_numeric = any(
            t in col_type
            for t in ["INT", "DOUBLE", "FLOAT", "DECIMAL", "NUMERIC", "BIGINT", "SMALLINT", "TINYINT"]
        )
        parse_ratio = _numeric_parse_ratio(base_sub, col_sql)
        mode = "histogram" if (pragma_numeric or parse_ratio >= 0.45) else "top_values"

    if mode == "threshold_3bins":
        if payload.x1 is None and payload.x2 is None:
            raise HTTPException(status_code=400, detail="threshold_3bins 至少需要提供 x1 或 x2")
        x1 = float(payload.x1 if payload.x1 is not None else payload.x2)
        x2 = float(payload.x2 if payload.x2 is not None else payload.x1)
        if x1 > x2:
            x1, x2 = x2, x1
        two_bins = abs(x1 - x2) < 1e-12
        finite_from = (
            f"SELECT try_cast(t.{col_sql} AS DOUBLE) AS v FROM ({{sub}}) t WHERE {fin_w}"
        )
        full_row = con.execute(
            f"""
            SELECT
              coalesce(sum(CASE WHEN v < ? THEN 1 ELSE 0 END),0)::BIGINT,
              coalesce(sum(CASE WHEN v >= ? AND v <= ? THEN 1 ELSE 0 END),0)::BIGINT,
              coalesce(sum(CASE WHEN v > ? THEN 1 ELSE 0 END),0)::BIGINT
            FROM ({finite_from.format(sub=base_sub)}) s
            """,
            [x1, x1, x2, x2],
        ).fetchone()
        filt_row = con.execute(
            f"""
            SELECT
              coalesce(sum(CASE WHEN v < ? THEN 1 ELSE 0 END),0)::BIGINT,
              coalesce(sum(CASE WHEN v >= ? AND v <= ? THEN 1 ELSE 0 END),0)::BIGINT,
              coalesce(sum(CASE WHEN v > ? THEN 1 ELSE 0 END),0)::BIGINT
            FROM ({finite_from.format(sub=filtered_sub)}) s
            """,
            [x1, x1, x2, x2],
        ).fetchone()
        labels = [f"(< {x1})", f"(>= {x1})"] if two_bins else [f"(< {x1})", f"[{x1}, {x2}]", f"(> {x2})"]
        items = []
        if two_bins:
            f0 = int(full_row[0]) if full_row and full_row[0] is not None else 0
            f1 = int(full_row[1]) if full_row and full_row[1] is not None else 0
            f2 = int(full_row[2]) if full_row and full_row[2] is not None else 0
            c0 = int(filt_row[0]) if filt_row and filt_row[0] is not None else 0
            c1 = int(filt_row[1]) if filt_row and filt_row[1] is not None else 0
            c2 = int(filt_row[2]) if filt_row and filt_row[2] is not None else 0
            pairs = [(labels[0], f0, c0), (labels[1], f1 + f2, c1 + c2)]
        else:
            pairs = []
            for idx, label in enumerate(labels):
                fv = int(full_row[idx]) if full_row and full_row[idx] is not None else 0
                cv = int(filt_row[idx]) if filt_row and filt_row[idx] is not None else 0
                pairs.append((label, fv, cv))
        for label, full_count, filtered_count in pairs:
            row: Dict[str, Any] = {"label": label, "count": filtered_count}
            if compare_effective:
                row["full_count"] = full_count
                row["filtered_count"] = filtered_count
            items.append(row)
        return {
            "table_name": table_name,
            "column": real_field,
            "mode": "threshold_3bins",
            "detected_type": "numeric",
            "items": items,
            "x1": x1,
            "x2": x2,
        }

    if mode == "top_values":
        top_n = int(payload.top_n)
        cat_where = f"WHERE {junk_w}"
        distinct_row = con.execute(
            f"""
            SELECT count(DISTINCT CAST(t.{col_sql} AS VARCHAR))
            FROM ({base_sub}) t
            {cat_where}
            """
        ).fetchone()
        distinct_cat = int(distinct_row[0]) if distinct_row and distinct_row[0] is not None else 0

        labels_rows = con.execute(
            f"""
            SELECT CAST(t.{col_sql} AS VARCHAR) AS label, count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            {cat_where}
            GROUP BY label
            ORDER BY cnt DESC
            LIMIT {top_n}
            """
        ).fetchall()
        labels = [(r[0] if r[0] is not None else "") for r in labels_rows]
        note: Optional[str] = None
        if distinct_cat > top_n:
            note = f"仅展示计数最高的前 {top_n} 个类别（去重共 {distinct_cat} 类）"
        if not labels:
            return {
                "table_name": table_name,
                "column": real_field,
                "mode": "top_values",
                "detected_type": "categorical",
                "items": [],
                "note": note,
            }
        in_clause = ",".join(sql_literal(v) for v in labels)
        full_rows = con.execute(
            f"""
            SELECT CAST(t.{col_sql} AS VARCHAR) AS label, count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            WHERE CAST(t.{col_sql} AS VARCHAR) IN ({in_clause}) AND {junk_w}
            GROUP BY label
            """
        ).fetchall()
        filt_rows = con.execute(
            f"""
            SELECT CAST(t.{col_sql} AS VARCHAR) AS label, count(*)::BIGINT AS cnt
            FROM ({filtered_sub}) t
            WHERE CAST(t.{col_sql} AS VARCHAR) IN ({in_clause}) AND {junk_w}
            GROUP BY label
            """
        ).fetchall()
        full_map = {("" if r[0] is None else str(r[0])): int(r[1]) for r in full_rows}
        filt_map = {("" if r[0] is None else str(r[0])): int(r[1]) for r in filt_rows}
        items = []
        for lb in labels:
            k = "" if lb is None else str(lb)
            shown = "(空)" if k == "" else k
            row = {"label": shown, "count": filt_map.get(k, 0)}
            if compare_effective:
                row["full_count"] = full_map.get(k, 0)
                row["filtered_count"] = filt_map.get(k, 0)
            items.append(row)
        out: Dict[str, Any] = {
            "table_name": table_name,
            "column": real_field,
            "mode": "top_values",
            "detected_type": "categorical",
            "items": items,
        }
        if note:
            out["note"] = note
        return out

    # histogram（含：有限数值、去重数≤bins 时离散展示、可选 bin_width）
    bins_req = max(3, min(200, int(payload.bins)))
    mm = con.execute(
        f"""
        SELECT
          min(try_cast(t.{col_sql} AS DOUBLE)) AS min_v,
          max(try_cast(t.{col_sql} AS DOUBLE)) AS max_v
        FROM ({base_sub}) t
        WHERE {fin_w}
        """
    ).fetchone()
    min_v = float(mm[0]) if mm and mm[0] is not None else None
    max_v = float(mm[1]) if mm and mm[1] is not None else None
    if min_v is None or max_v is None:
        return {
            "table_name": table_name,
            "column": real_field,
            "mode": "histogram",
            "detected_type": "numeric",
            "items": [],
        }

    drow = con.execute(
        f"""
        SELECT count(DISTINCT try_cast(t.{col_sql} AS DOUBLE))
        FROM ({base_sub}) t
        WHERE {fin_w}
        """
    ).fetchone()
    distinct_vals = int(drow[0]) if drow and drow[0] is not None else 0

    # 若用户显式指定了步长，则优先按步长分桶；仅在未指定步长时走离散值分布捷径
    if max_v == min_v or (payload.bin_width is None and distinct_vals <= bins_req):
        full_rows = con.execute(
            f"""
            SELECT try_cast(t.{col_sql} AS DOUBLE) AS v, count(*)::BIGINT AS cnt
            FROM ({base_sub}) t
            WHERE {fin_w}
            GROUP BY v
            ORDER BY v
            """
        ).fetchall()
        filt_rows = con.execute(
            f"""
            SELECT try_cast(t.{col_sql} AS DOUBLE) AS v, count(*)::BIGINT AS cnt
            FROM ({filtered_sub}) t
            WHERE {fin_w}
            GROUP BY v
            ORDER BY v
            """
        ).fetchall()
        full_map_d = {float(r[0]): int(r[1]) for r in full_rows if r[0] is not None}
        filt_map_d = {float(r[0]): int(r[1]) for r in filt_rows if r[0] is not None}
        items_d: List[Dict[str, Any]] = []
        for v in sorted(set(full_map_d.keys()) | set(filt_map_d.keys())):
            fv = full_map_d.get(v, 0)
            cv = filt_map_d.get(v, 0)
            lbl = str(int(v)) if abs(v - round(v)) < 1e-9 else f"{v:.6g}"
            row = {"label": lbl, "count": cv}
            if compare_effective:
                row["full_count"] = fv
                row["filtered_count"] = cv
            items_d.append(row)
        return {
            "table_name": table_name,
            "column": real_field,
            "mode": "histogram",
            "detected_type": "numeric",
            "histogram_shape": "discrete_values",
            "items": items_d,
        }

    span = max_v - min_v
    if payload.bin_width is not None and float(payload.bin_width) > 0:
        bw = float(payload.bin_width)
        anchor = math.floor(min_v / bw) * bw
        bins_eff = max(1, min(200, int(math.ceil((max_v - anchor) / bw)) + 1))
    else:
        bins_eff = bins_req
        bw = span / bins_eff
        anchor = math.floor(min_v)

    def bucket_counts(sub_sql: str) -> Dict[int, int]:
        rs = con.execute(
            f"""
            WITH s AS (
              SELECT try_cast(t.{col_sql} AS DOUBLE) AS v
              FROM ({sub_sql}) t
              WHERE {fin_w}
            )
            SELECT LEAST(FLOOR((v - ?) / ?), ?)::INT AS bucket, count(*)::BIGINT AS cnt
            FROM s
            GROUP BY bucket
            """,
            [anchor, bw, bins_eff - 1],
        ).fetchall()
        return {int(r[0]): int(r[1]) for r in rs}

    full_map = bucket_counts(base_sub)
    filt_map = bucket_counts(filtered_sub)
    items = []
    for b in range(bins_eff):
        left = anchor + b * bw
        right = left + bw
        row = {
            "label": f"[{pretty_num(left)}, {pretty_num(right)}{']' if b == bins_eff - 1 else ')'}",
            "count": filt_map.get(b, 0),
        }
        if compare_effective:
            row["full_count"] = full_map.get(b, 0)
            row["filtered_count"] = filt_map.get(b, 0)
        items.append(row)
    out_hist: Dict[str, Any] = {
        "table_name": table_name,
        "column": real_field,
        "mode": "histogram",
        "detected_type": "numeric",
        "bins": bins_eff,
        "bin_width_used": bw,
        "items": items,
    }
    if payload.bin_width is not None:
        out_hist["bin_width_requested"] = float(payload.bin_width)
    return out_hist


@app.post("/api/distribution")
def api_distribution(payload: UnifiedDistributionRequest) -> Dict[str, Any]:
    return _distribution_core(payload)


@app.post("/distribution")
def distribution(payload: UnifiedDistributionRequest) -> Dict[str, Any]:
    return _distribution_core(payload)


FRONTEND_INDEX = FRONTEND_DIR / "index.html"

if (FRONTEND_DIR / "static").exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR / "static")), name="frontend-static")


@app.get("/")
def frontend_root() -> FileResponse:
    if FRONTEND_INDEX.exists():
        return FileResponse(str(FRONTEND_INDEX))
    raise HTTPException(status_code=404, detail="前端资源未构建，请先执行前端 build")


@app.get("/{full_path:path}")
def frontend_fallback(full_path: str) -> FileResponse:
    api_prefixes = (
        "upload",
        "upload_pasted",
        "preview",
        "preview_result",
        "preview_station_traffic",
        "preview_station_traffic_column_distinct",
        "preview_station_traffic_distinct_count",
        "station_distinct_count",
        "condition_chart_counts",
        "filter",
        "nl_filter",
        "nl_filter_parse",
        "export",
        "export_nl_batch",
        "export_filtered_preview",
        "engineering_chart",
        "column_distribution",
        "distribution_compare",
        "distribution",
        "api",
        "column_distinct",
        "field_max",
        "health",
        "docs",
        "openapi.json",
        "redoc",
        "static",
    )
    if full_path.startswith(api_prefixes):
        raise HTTPException(status_code=404, detail="Not Found")
    if FRONTEND_INDEX.exists():
        return FileResponse(str(FRONTEND_INDEX))
    raise HTTPException(status_code=404, detail="前端资源未构建，请先执行前端 build")
